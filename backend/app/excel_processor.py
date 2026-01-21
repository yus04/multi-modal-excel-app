import io
import os
import base64
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import logging

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Process Excel files to extract content and images"""
    
    @staticmethod
    def extract_images_from_excel(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
        """Extract all images from an Excel file"""
        images = []
        
        try:
            # Use BytesIO to load the Excel file directly from bytes
            file_stream = io.BytesIO(file_content)
            workbook = load_workbook(file_stream)
            
            for sheet_idx, sheet in enumerate(workbook.worksheets):
                if hasattr(sheet, '_images'):
                    for img_idx, img in enumerate(sheet._images):
                        try:
                            image_data = img.ref
                            if hasattr(img, '_data'):
                                pil_image = Image.open(io.BytesIO(img._data()))
                            elif hasattr(image_data, 'data'):
                                pil_image = Image.open(io.BytesIO(image_data.data))
                            else:
                                continue
                            
                            # Convert to PNG and encode as base64
                            img_byte_arr = io.BytesIO()
                            pil_image.save(img_byte_arr, format='PNG')
                            img_byte_arr.seek(0)
                            
                            img_base64 = base64.b64encode(img_byte_arr.read()).decode('utf-8')
                            
                            # Get anchor position if available
                            anchor = getattr(img, 'anchor', None)
                            position = None
                            if anchor and hasattr(anchor, '_from'):
                                position = {
                                    'col': anchor._from.col,
                                    'row': anchor._from.row
                                }
                            
                            images.append({
                                'sheet': sheet.title,
                                'sheet_index': sheet_idx,
                                'image_index': img_idx,
                                'data': img_base64,
                                'position': position,
                                'filename': f"{os.path.splitext(filename)[0]}_sheet{sheet_idx}_img{img_idx}.png"
                            })
                        except Exception as e:
                            logger.warning(f"Error extracting image {img_idx} from sheet {sheet.title}: {str(e)}")
            
            return images
            
        except Exception as e:
            logger.error(f"Error processing Excel file {filename}: {str(e)}")
            raise
    
    @staticmethod
    def extract_text_from_excel(file_content: bytes) -> List[Dict[str, Any]]:
        """Extract text content from Excel sheets with row numbers"""
        sheets_data = []
        
        try:
            # Use BytesIO to load the Excel file directly from bytes
            file_stream = io.BytesIO(file_content)
            workbook = load_workbook(file_stream, data_only=True)
            
            for sheet_idx, sheet in enumerate(workbook.worksheets):
                rows_data = []
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    # Filter out empty rows
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    if any(val.strip() for val in row_values):
                        rows_data.append({
                            'row_number': row_idx,
                            'values': row_values
                        })
                
                sheets_data.append({
                    'sheet_name': sheet.title,
                    'sheet_index': sheet_idx,
                    'rows': rows_data
                })
            
            return sheets_data
            
        except Exception as e:
            logger.error(f"Error extracting text from Excel: {str(e)}")
            raise
